"""
ColourExtract.py

Scans an Excel/ODS workbook for cells with red or yellow fill (or bold text).
Orange cells are ignored.

Output rules:
  - Single merged output sheet across all input sheets
  - Row 6 of each input sheet provides the dataPoint column headers
  - For each marked cell: title (col A of that row) + cell value under its dataPoint column
  - Red cells are coloured red in output; yellow cells are coloured yellow
  - Multiple cells with the same dataPoint share the same output column

Supported input formats : .ods, .xlsx, .xlsm
Output format           : .xlsx

========== SETUP (run once in PowerShell) ==========

1. Install Python 3.14 if not already present:
      winget install Python.Python.3.14

2. Confirm Python is available:
      py --version

3. Upgrade pip:
      py -m pip install --upgrade pip

4. Install required libraries:
      py -m pip install openpyxl odfpy

========== USAGE ==========

    py ColourExtract.py <input.ods or input.xlsx> <output.xlsx>

Example:
    py ColourExtract.py "C:\data\MySheet.ods" "C:\data\output.xlsx"
"""

import os
import sys

# Row (1-based) that holds the dataPoint column headers in each input sheet
DATAPOINT_HEADER_ROW = 6

# ---------------------------------------------------------------------------
# Colour classification
# ---------------------------------------------------------------------------

RED_RGB    = {"ff0000", "c00000", "ff4444", "cc0000", "ff0000"}
YELLOW_RGB = {"ffff00", "ffd966", "ffc000", "ffe699", "fff2cc", "ffff99", "ffcc00"}
ORANGE_RGB = {"ff6600", "ff8000", "ffa500", "ff7f00", "ffb347", "ff8c00", "ff6a00",
              "ff4500", "ff7722", "e25822"}


def normalise_colour(c):
    if not c:
        return ""
    c = c.lower().strip().lstrip("#")
    if len(c) == 8:  # aarrggbb -> rrggbb
        c = c[2:]
    return c


def classify_fill(bg):
    """Returns 'red', 'yellow', 'orange', or None."""
    n = normalise_colour(bg or "")
    if n in RED_RGB:
        return "red"
    if n in YELLOW_RGB:
        return "yellow"
    if n in ORANGE_RGB:
        return "orange"
    return None


# ---------------------------------------------------------------------------
# ODS reader
# ---------------------------------------------------------------------------


def read_ods(path):
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = load(path)
    style_map = {}

    def parse_styles(container):
        if not container:
            return
        for node in container.childNodes:
            if not hasattr(node, "getAttribute"):
                continue
            try:
                name = node.getAttribute("name")
            except (ValueError, AttributeError):
                continue
            if not name:
                continue
            bg, bold = None, False
            for ch in node.childNodes:
                if not hasattr(ch, "qname"):
                    continue
                loc = ch.qname[1]
                if loc == "table-cell-properties":
                    try:
                        raw = (
                            ch.getAttribute("backgroundcolor")
                            or ch.getAttribute("background-color")
                            or ""
                        )
                    except (ValueError, AttributeError):
                        raw = ""
                    if raw and raw.lower() not in ("", "transparent", "#00000000"):
                        bg = raw
                elif loc == "text-properties":
                    try:
                        fw = ch.getAttribute("fontweight") or ch.getAttribute("font-weight") or ""
                    except (ValueError, AttributeError):
                        fw = ""
                    if fw.lower() == "bold":
                        bold = True
            style_map[name] = {"bg": bg, "bold": bold}

    parse_styles(doc.automaticstyles)
    parse_styles(doc.styles)

    result = {}
    for table in doc.spreadsheet.getElementsByType(Table):
        sname = table.getAttribute("name")
        srows = []
        for tr in table.getElementsByType(TableRow):
            row_repeat = int(tr.getAttribute("numberrowsrepeated") or 1)
            rcells = []
            for cell in tr.getElementsByType(TableCell):
                rep = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                si = style_map.get(cell.getAttribute("stylename") or "", {"bg": None, "bold": False})
                vt = cell.getAttribute("valuetype") or ""
                if vt == "float":
                    try:
                        n = float(cell.getAttribute("value"))
                        v = int(n) if n.is_integer() else n
                    except (TypeError, ValueError):
                        v = None
                elif vt == "date":
                    v = cell.getAttribute("datevalue")
                else:
                    ps = cell.getElementsByType(P)
                    t = "".join(
                        nd.data for p in ps for nd in p.childNodes if hasattr(nd, "data")
                    )
                    v = t if t else None
                ci = {"value": v, "bg": si["bg"], "bold": si["bold"]}
                for _ in range(min(rep, 50)):
                    rcells.append(ci)
            # Expand repeated rows so indices match what the user sees in the spreadsheet
            for _ in range(min(row_repeat, 200)):
                srows.append(rcells)
        result[sname] = srows
    return result


# ---------------------------------------------------------------------------
# XLSX reader
# ---------------------------------------------------------------------------


def read_xlsx(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    result = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        srows = []
        for row in ws.iter_rows():
            rc = []
            for cell in row:
                fill = cell.fill
                bg = None
                if fill and fill.fill_type not in (None, "none"):
                    fg = fill.fgColor
                    if fg and fg.type == "rgb":
                        bg = fg.rgb
                rc.append({"value": cell.value, "bg": bg, "bold": bool(cell.font and cell.font.bold)})
            srows.append(rc)
        result[sname] = srows
    return result


# ---------------------------------------------------------------------------
# Core extraction  (shared by both readers)
# ---------------------------------------------------------------------------


def extract_all_sheets(sheets_data):
    """
    sheets_data: {sheet_name: [row, ...]}  where row = [cell_info, ...]
    cell_info = {value, bg, bold}

    Returns:
        entries      - list of {title, datapoint, value, colour}
        datapoints   - ordered list of unique dataPoint names (preserves first-seen order)
    """
    entries = []
    datapoints = []          # ordered unique dataPoint names
    dp_seen = set()

    for sname, rows in sheets_data.items():
        # Row 6 (1-based) → index 5: dataPoint headers
        if len(rows) < DATAPOINT_HEADER_ROW:
            print(f"  Sheet '{sname}': fewer than {DATAPOINT_HEADER_ROW} rows, skipped")
            continue

        header_row = rows[DATAPOINT_HEADER_ROW - 1]
        # col_index -> dataPoint name  (skip col 0 = title column)
        col_to_dp = {}
        for ci, cell in enumerate(header_row):
            if ci == 0:
                continue
            dp_name = str(cell["value"]).strip() if cell["value"] is not None else ""
            if dp_name:
                col_to_dp[ci] = dp_name
                if dp_name not in dp_seen:
                    dp_seen.add(dp_name)
                    datapoints.append(dp_name)

        found = 0
        for row_idx, row in enumerate(rows):
            if row_idx == DATAPOINT_HEADER_ROW - 1:  # skip the header row itself
                continue
            if not row:
                continue

            title = row[0]["value"] if row else None

            for ci, cell in enumerate(row):
                if ci == 0:
                    continue  # title column is never a data cell
                if ci not in col_to_dp:
                    continue  # column has no dataPoint header

                colour = classify_fill(cell.get("bg"))
                if colour != "red" and colour != "yellow":
                    continue  # only red and yellow cells are processed

                entries.append({
                    "title": title,
                    "datapoint": col_to_dp[ci],
                    "value": cell["value"],
                    "colour": colour,   # 'red' | 'yellow' | 'bold'
                })
                found += 1

        print(f"  Sheet '{sname}': {found} marked cell(s) extracted")

    return entries, datapoints


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------


def write_xlsx(entries, datapoints, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    RED_FILL    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    HEADER_FONT = Font(bold=True)

    # Merge entries that share the same title into one row, preserving first-seen order
    merged = {}       # title -> {datapoint -> {value, colour}}
    title_order = []  # keeps insertion order
    for entry in entries:
        t = entry["title"]
        if t not in merged:
            merged[t] = {}
            title_order.append(t)
        merged[t][entry["datapoint"]] = {"value": entry["value"], "colour": entry["colour"]}

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Title"] + datapoints
    dp_col = {dp: i + 2 for i, dp in enumerate(datapoints)}

    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = HEADER_FONT

    for row_idx, title in enumerate(title_order, 2):
        ws.cell(row_idx, 1, title)
        for dp, info in merged[title].items():
            cell = ws.cell(row_idx, dp_col[dp], info["value"])
            if info["colour"] == "red":
                cell.fill = RED_FILL
            elif info["colour"] == "yellow":
                cell.fill = YELLOW_FILL

    wb.save(path)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

READERS = {".ods": read_ods, ".xlsx": read_xlsx, ".xlsm": read_xlsx, ".xltx": read_xlsx}


def process(inp, out):
    ie = os.path.splitext(inp)[1].lower()
    oe = os.path.splitext(out)[1].lower()
    if ie not in READERS:
        print(f"Unsupported input format: {ie}  (supported: {', '.join(READERS)})")
        sys.exit(1)
    if oe != ".xlsx":
        print("Output format must be .xlsx")
        sys.exit(1)

    print(f"Reading: {inp}")
    sheets = READERS[ie](inp)

    entries, datapoints = extract_all_sheets(sheets)

    if not entries:
        print("No marked cells found. Output not written.")
        return

    write_xlsx(entries, datapoints, out)
    print(f"Written: {out}  ({len(entries)} entry/entries, {len(datapoints)} dataPoint column(s))")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: py ColourExtract.py <input.ods/.xlsx> <output.ods/.xlsx>")
        sys.exit(1)
    process(sys.argv[1], sys.argv[2])
